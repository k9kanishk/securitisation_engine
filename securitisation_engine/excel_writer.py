from __future__ import annotations

from typing import Dict

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name: str) -> str:
    name = str(name)
    return name[:31]


def _write_df(ws, df: pd.DataFrame, start_row: int, start_col: int, header: bool = True, number_format: str = "#,##0.00"):
    bold = Font(bold=True)
    align = Alignment(vertical="top")

    r = start_row
    c = start_col

    if header:
        for j, col_name in enumerate(df.columns, start=c):
            cell = ws.cell(row=r, column=j, value=str(col_name))
            cell.font = bold
            cell.alignment = align
        r += 1

    for i in range(len(df)):
        for j, col_name in enumerate(df.columns, start=c):
            val = df.iloc[i][col_name]
            cell = ws.cell(row=r + i, column=j, value=val)
            cell.alignment = align
            if isinstance(val, (int, float)) and col_name not in ("Payment Date", "metric", "value"):
                cell.number_format = number_format

    # autosize columns (lightweight)
    for j in range(c, c + len(df.columns)):
        col_letter = get_column_letter(j)
        max_len = 10
        for rr in range(start_row, start_row + 1 + len(df)):
            v = ws.cell(row=rr, column=j).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)) if len(str(v)) < 60 else 60)
        ws.column_dimensions[col_letter].width = max_len + 2


def ensure_template(path: str) -> None:
    """
    Creates a clean IPD template if it doesn't exist.
    Sheets:
      - Available Funds
      - Priority of Payments - Interest
      - Priority of Payments - Principal
      - Note Rollforward
      - Investor Summary
    """
    wb = Workbook()
    wb.remove(wb.active)

    for name in [
        "Available Funds",
        "Priority of Payments - Interest",
        "Priority of Payments - Principal",
        "Note Rollforward",
        "Investor Summary",
    ]:
        wb.create_sheet(_safe_sheet_name(name))

    wb.save(path)


def write_ipd_pack(
    template_path: str,
    output_path: str,
    dfs: Dict[str, pd.DataFrame],
) -> None:
    wb = load_workbook(template_path)

    for sheet_name, df in dfs.items():
        safe = _safe_sheet_name(sheet_name)
        if safe not in wb.sheetnames:
            wb.create_sheet(safe)
        ws = wb[safe]
        ws.delete_rows(1, ws.max_row)  # clear
        _write_df(ws, df, start_row=1, start_col=1, header=True)

    wb.save(output_path)
